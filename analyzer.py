#!/usr/bin/env python3
"""
Particle Diameter Analyzer
==========================
Identifies and measures circular particles in microscopy images.

Usage:
    python particle_analyzer.py <image_path> <scale_bar_um> <scale_bar_pixels>

Example:
    python particle_analyzer.py sample.bmp 100 150
    # This means the 100 μm scale bar is 150 pixels in the image
    # Output: sample_analyzed.xlsx with 'image' and 'measurements' sheets

Outputs:
    - <image_name>_analyzed.xlsx : Excel file with two sheets:
        - 'image' : Annotated image with detected particles
        - 'measurements' : CSV data with all particle measurements
"""

import cv2
import numpy as np
import pandas as pd
from scipy import ndimage
from skimage.segmentation import watershed
from skimage.feature import peak_local_max
import argparse
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io


def detect_particles(gray_image):
    """
    Detect particles using adaptive threshold + watershed segmentation.
    
    Returns list of particle dictionaries with measurements.
    """
    h, w = gray_image.shape
    
    # Adaptive thresholding
    blurred = cv2.GaussianBlur(gray_image, (5, 5), 0)
    adaptive = cv2.adaptiveThreshold(
        blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, blockSize=21, C=6
    )
    
    # Fill holes to handle hollow particles
    mask = np.zeros((h + 2, w + 2), np.uint8)
    filled = adaptive.copy()
    cv2.floodFill(filled, mask, (0, 0), 255)
    filled_inv = cv2.bitwise_not(filled)
    filled_particles = adaptive | filled_inv
    
    # Morphological cleanup
    kernel = np.ones((3, 3), np.uint8)
    filled_particles = cv2.morphologyEx(filled_particles, cv2.MORPH_OPEN, kernel, iterations=2)
    filled_particles = cv2.morphologyEx(filled_particles, cv2.MORPH_CLOSE, kernel, iterations=1)
    
    # Distance transform for watershed
    dist_transform = ndimage.distance_transform_edt(filled_particles)
    
    # Find local maxima (particle centers)
    coords = peak_local_max(
        dist_transform,
        min_distance=4,
        threshold_rel=0.15,
        labels=filled_particles,
        exclude_border=False
    )
    
    # Create markers for watershed
    markers = np.zeros(filled_particles.shape, dtype=np.int32)
    for i, (y, x) in enumerate(coords):
        markers[y, x] = i + 1
    markers = ndimage.grey_dilation(markers, size=(3, 3))
    
    # Watershed segmentation
    labels = watershed(-dist_transform, markers, mask=filled_particles)
    
    # First pass: collect particles with good circularity to establish baseline
    good_circ_particles = []
    for label_id in range(1, labels.max() + 1):
        mask_single = (labels == label_id).astype(np.uint8)
        area = np.sum(mask_single)
        
        if area < 50:
            continue
        
        contours, _ = cv2.findContours(mask_single, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            continue
        
        cnt = contours[0]
        (cx, cy), radius = cv2.minEnclosingCircle(cnt)
        perimeter = cv2.arcLength(cnt, True)
        circularity = 4 * np.pi * area / (perimeter ** 2) if perimeter > 0 else 0
        
        if circularity > 0.75:
            good_circ_particles.append({'area': area, 'radius': radius, 'circularity': circularity})
    
    # Calculate size statistics
    median_area = np.median([p['area'] for p in good_circ_particles]) if good_circ_particles else 400
    q75_area = np.percentile([p['area'] for p in good_circ_particles], 75) if good_circ_particles else 600
    max_allowed_area = q75_area * 2.5
    
    # Second pass: filter with size awareness
    particles = []
    for label_id in range(1, labels.max() + 1):
        mask_single = (labels == label_id).astype(np.uint8)
        area = np.sum(mask_single)
        
        if area < 50 or area > max_allowed_area:
            continue
        
        contours, _ = cv2.findContours(mask_single, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            continue
        
        cnt = contours[0]
        (cx, cy), radius = cv2.minEnclosingCircle(cnt)
        perimeter = cv2.arcLength(cnt, True)
        circularity = 4 * np.pi * area / (perimeter ** 2) if perimeter > 0 else 0
        enclosing_area = np.pi * radius ** 2
        fill_ratio = area / enclosing_area if enclosing_area > 0 else 0
        
        # Size-dependent circularity threshold
        if area > q75_area * 1.5:
            min_circ = 0.65
        elif area > median_area * 1.5:
            min_circ = 0.55
        else:
            min_circ = 0.45
        
        if circularity < min_circ:
            continue
        
        # Hollow detection: compare center brightness vs ring brightness
        cx_int, cy_int = int(cx), int(cy)
        r_center = max(1, int(radius * 0.25))
        y1c, y2c = max(0, cy_int - r_center), min(h, cy_int + r_center + 1)
        x1c, x2c = max(0, cx_int - r_center), min(w, cx_int + r_center + 1)
        center_val = np.mean(gray_image[y1c:y2c, x1c:x2c]) if (y2c > y1c and x2c > x1c) else 0
        
        ring_vals = []
        r_int = max(3, int(radius))
        for angle in range(0, 360, 30):
            for r_frac in [0.6, 0.75]:
                rx = int(cx + r_int * r_frac * np.cos(np.radians(angle)))
                ry = int(cy + r_int * r_frac * np.sin(np.radians(angle)))
                if 0 <= rx < w and 0 <= ry < h:
                    ring_vals.append(gray_image[ry, rx])
        ring_val = np.mean(ring_vals) if ring_vals else center_val
        is_hollow = (center_val - ring_val) > 15
        
        particles.append({
            'center_x': cx,
            'center_y': cy,
            'radius': radius,
            'diameter_px': 2 * radius,
            'area_px': area,
            'circularity': circularity,
            'fill_ratio': fill_ratio,
            'is_hollow': is_hollow
        })
    
    return particles


def create_annotated_image(original_image, particles):
    """
    Create annotated image with detected particles marked.
    
    Colors:
        - Yellow: Hollow particles
        - Green: Filled particles
        - Red dot: Center point
        - Blue text: Particle ID
    """
    vis = original_image.copy()
    
    for i, p in enumerate(particles):
        cx, cy = int(p['center_x']), int(p['center_y'])
        r = int(p['radius'])
        particle_id = i + 1
        
        # Circle color based on hollow/filled
        color = (0, 255, 255) if p['is_hollow'] else (0, 255, 0)  # Yellow or Green (BGR)
        
        # Draw circle and center
        cv2.circle(vis, (cx, cy), r, color, 2)
        cv2.circle(vis, (cx, cy), 2, (0, 0, 255), -1)  # Red center
        
        # Add ID label
        cv2.putText(vis, str(particle_id), (cx + r + 2, cy + 4),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.35, (255, 0, 0), 1)  # Blue text
    
    return vis


def create_measurements_csv(particles, um_per_pixel):
    """
    Create DataFrame with particle measurements.
    """
    records = []
    for i, p in enumerate(particles):
        records.append({
            'particle_id': i + 1,
            'center_x_px': round(p['center_x'], 1),
            'center_y_px': round(p['center_y'], 1),
            'diameter_px': round(p['diameter_px'], 2),
            'diameter_um': round(p['diameter_px'] * um_per_pixel, 2),
            'area_px': int(p['area_px']),
            'circularity': round(p['circularity'], 3),
            'is_hollow': p['is_hollow']
        })
    
    return pd.DataFrame(records)


def analyze_image(image_path, scale_bar_um, scale_bar_pixels, output_dir=None):
    """
    Main analysis function.
    
    Args:
        image_path: Path to input image
        scale_bar_um: Scale bar length in micrometers
        scale_bar_pixels: Scale bar length in pixels
        output_dir: Output directory (default: same as input)
    
    Returns:
        DataFrame with measurements
    """
    # Calculate scale
    um_per_pixel = scale_bar_um / scale_bar_pixels
    
    # Load image
    img = cv2.imread(str(image_path))
    if img is None:
        raise ValueError(f"Could not load image: {image_path}")
    
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    print(f"Image size: {img.shape[1]} x {img.shape[0]} pixels")
    print(f"Scale: {um_per_pixel:.4f} μm/pixel ({scale_bar_um} μm = {scale_bar_pixels} px)")
    
    # Detect particles
    print("\nDetecting particles...")
    particles = detect_particles(gray)
    print(f"Detected {len(particles)} particles")
    
    # Statistics
    hollow_count = sum(1 for p in particles if p['is_hollow'])
    diameters_um = [p['diameter_px'] * um_per_pixel for p in particles]
    
    print(f"\nStatistics:")
    print(f"  Hollow particles: {hollow_count}")
    print(f"  Filled particles: {len(particles) - hollow_count}")
    print(f"  Diameter range: {min(diameters_um):.1f} - {max(diameters_um):.1f} μm")
    print(f"  Mean diameter: {np.mean(diameters_um):.1f} μm")
    print(f"  Median diameter: {np.median(diameters_um):.1f} μm")
    
    # Create outputs
    input_path = Path(image_path)
    if output_dir is None:
        output_dir = input_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    
    base_name = input_path.stem
    
    # Create annotated image
    annotated = create_annotated_image(img, particles)
    
    # Create DataFrame with measurements
    df = create_measurements_csv(particles, um_per_pixel)
    
    # Save to Excel file with two sheets
    excel_path = output_dir / f"{base_name}_analyzed.xlsx"
    
    # Write DataFrame to Excel first (creates workbook)
    with pd.ExcelWriter(str(excel_path), engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='measurements', index=False)
    
    # Open the workbook to add image sheet
    wb = load_workbook(str(excel_path))
    
    # Create 'image' sheet and add image
    ws_image = wb.create_sheet("image")
    
    # Convert BGR to RGB for PIL
    annotated_rgb = cv2.cvtColor(annotated, cv2.COLOR_BGR2RGB)
    pil_image = Image.fromarray(annotated_rgb)
    
    # Save PIL image to bytes
    img_bytes = io.BytesIO()
    pil_image.save(img_bytes, format='PNG')
    img_bytes.seek(0)
    
    # Add image to worksheet
    xl_img = XLImage(img_bytes)
    # Scale image to fit (optional: adjust size as needed)
    # Limit image size to reasonable dimensions for Excel
    max_width = 1200
    max_height = 800
    if xl_img.width > max_width or xl_img.height > max_height:
        scale = min(max_width / xl_img.width, max_height / xl_img.height)
        xl_img.width = int(xl_img.width * scale)
        xl_img.height = int(xl_img.height * scale)
    
    ws_image.add_image(xl_img, 'A1')
    
    # Adjust column width for image sheet (optional)
    ws_image.column_dimensions['A'].width = max(10, xl_img.width / 7)  # Approximate
    
    # Save workbook
    wb.save(str(excel_path))
    print(f"\nSaved Excel file with image and measurements: {excel_path}")
    
    return df


def main():
    parser = argparse.ArgumentParser(
        description='Analyze particle diameters in microscopy images.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python particle_analyzer.py image.bmp 100 150
        Analyze image.bmp where the 100 μm scale bar is 150 pixels

    python particle_analyzer.py image.bmp 50 75 --output ./results
        Save outputs to ./results directory
        """
    )
    
    parser.add_argument('image_path', help='Path to input image')
    parser.add_argument('scale_bar_um', type=float, help='Scale bar length in micrometers')
    parser.add_argument('scale_bar_pixels', type=float, help='Scale bar length in pixels')
    parser.add_argument('--output', '-o', help='Output directory (default: same as input)')
    
    args = parser.parse_args()
    
    analyze_image(args.image_path, args.scale_bar_um, args.scale_bar_pixels, args.output)


if __name__ == '__main__':
    main()